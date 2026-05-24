"""生成团队名册.xlsx — 铁律量化项目成员名册 Excel 版本"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# --- Styles ---
DARK_BG = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
MID_BG = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
LIGHT_BG = PatternFill(start_color="F0F2F5", end_color="F0F2F5", fill_type="solid")
WHITE_BG = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
RED_FONT = Font(name="Microsoft YaHei", bold=True, color="E74C3C", size=11)
GREEN_FONT = Font(name="Microsoft YaHei", bold=True, color="27AE60", size=11)
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Microsoft YaHei", bold=True, color="1A1A2E", size=16)
SUBTITLE_FONT = Font(name="Microsoft YaHei", bold=True, color="1A1A2E", size=12)
NORMAL_FONT = Font(name="Microsoft YaHei", color="333333", size=10)
SMALL_FONT = Font(name="Microsoft YaHei", color="666666", size=9)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

def style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = DARK_BG
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_data_cell(ws, row, col, align=CENTER):
    cell = ws.cell(row=row, column=col)
    cell.font = NORMAL_FONT
    cell.alignment = align
    cell.border = THIN_BORDER
    cell.fill = WHITE_BG
    return cell

def auto_width(ws, min_w=8, max_w=50):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split('\n')
                max_line = max(len(line) for line in lines)
                # Approximate: CJK chars ~2, ASCII ~1
                cjk = sum(1 for ch in str(cell.value) if '一' <= ch <= '鿿')
                ascii_chars = len(str(cell.value)) - cjk
                lengths.append(cjk * 2.2 + ascii_chars * 1.1)
        if lengths:
            w = min(max(max(lengths) + 2, min_w), max_w)
            ws.column_dimensions[col_letter].width = w

today = date.today().isoformat()

wb = openpyxl.Workbook()

# ============ Sheet 1: 成员总览 ============
ws1 = wb.active
ws1.title = "成员总览"

# Title
ws1.merge_cells("A1:G1")
ws1.cell(row=1, column=1, value="铁律量化 — 团队名册").font = TITLE_FONT
ws1.cell(row=1, column=1).alignment = CENTER

ws1.merge_cells("A2:G2")
ws1.cell(row=2, column=1, value=f"版本 v1.8 | {today} | 项目总监 阿黑").font = Font(name="Microsoft YaHei", color="666666", size=10)
ws1.cell(row=2, column=1).alignment = CENTER

# Architecture diagram
ws1.merge_cells("A4:G4")
ws1.cell(row=4, column=1, value="团队架构 — 三线并行，各司其职").font = SUBTITLE_FONT

ws1.merge_cells("A5:G5")
ws1.cell(row=5, column=1, value="金融线: 用户(创始人) → 阿黑(项目总监) → 腰子(金融负责人) ← 山猫 | 玉夜 | 流金 | 青山 (专业支撑)").font = Font(name="Microsoft YaHei", color="1A1A2E", size=10)
ws1.cell(row=5, column=1).alignment = CENTER
ws1.cell(row=5, column=1).fill = LIGHT_BG

ws1.merge_cells("A6:G6")
ws1.cell(row=6, column=1, value="工程线: 用户(创始人) → 阿黑(项目总监) → 情墨 | 千光 | 红枫 | 新安 | 红结").font = Font(name="Microsoft YaHei", color="16213E", size=10)
ws1.cell(row=6, column=1).alignment = CENTER
ws1.cell(row=6, column=1).fill = LIGHT_BG

ws1.merge_cells("A7:G7")
ws1.cell(row=7, column=1, value="审计线: 旧影(审计官) → 阿黑(项目总监)  [独立于金融线和工程线，直接对阿黑负责]").font = Font(name="Microsoft YaHei", color="E67E22", size=10)
ws1.cell(row=7, column=1).alignment = CENTER
ws1.cell(row=7, column=1).fill = LIGHT_BG

# Member overview table
headers = ["角色", "名称", "召唤", "核心职责", "一句话", "完整定义", "召唤命令"]
row = 9
for c, h in enumerate(headers, 1):
    ws1.cell(row=row, column=c, value=h)
style_header_row(ws1, row, len(headers))

members = [
    ["项目总监", "阿黑", "—", "调度团队、跟踪全局、主动运营", "让合适的人做合适的事",
     ".claude/agents/项目总监-阿黑.md", "—"],
    ["金融业务负责人", "腰子", "/腰子", "投资分析决策核心，协调四位专家，对金融判断质量负责", "告诉你该不该买，买多少，什么时候卖",
     ".claude/agents/金融专家-腰子.md", ".claude/commands/腰子.md"],
    ["宏观巡检官", "山猫", "/山猫", "政策信号、经济日历、市场情绪 → 支撑腰子宏观判断", "告诉腰子外部环境怎么变",
     ".claude/agents/宏观巡检-山猫.md", ".claude/commands/山猫.md"],
    ["数据监理", "玉夜", "/玉夜", "API巡检、缓存审计、数据完整性 → 支撑腰子数据信心", "告诉腰子数据能不能用",
     ".claude/agents/数据监理-玉夜.md", ".claude/commands/玉夜.md"],
    ["风控官", "流金", "/流金", "风险审计、过拟合检测、交易纪律 → 支撑腰子风险约束", "告诉腰子风险边界在哪里",
     ".claude/agents/风控官-流金.md", ".claude/commands/流金.md"],
    ["策略研究员", "青山", "/青山", "因子有效性、策略优化、规律挖掘 → 支撑腰子策略进化", "告诉腰子怎么让策略更聪明",
     ".claude/agents/策略研究员-青山.md", ".claude/commands/青山.md"],
    ["系统架构师", "情墨", "/情墨", "模块设计、技术选型、API契约、重构决策", "告诉你系统该怎么搭",
     ".claude/agents/系统架构师-情墨.md", ".claude/commands/情墨.md"],
    ["构建工程师(CI)", "千光", "/千光", "自动化流水线、定时调度、代码生成", "让脚本自己跑起来",
     ".claude/agents/构建工程师-千光.md", ".claude/commands/千光.md"],
    ["部署工程师(环境)", "红枫", "/红枫", "环境配置、依赖管理、版本发布、运行监控", "让系统跑得稳",
     ".claude/agents/部署工程师-红枫.md", ".claude/commands/红枫.md"],
    ["质量工程师(测试)", "新安", "/新安", "测试策略、回归验证、变更影响分析", "证明改了不会坏",
     ".claude/agents/质量工程师-新安.md", ".claude/commands/新安.md"],
    ["代码工匠", "红结", "/红结", "PowerShell/Python编码实现、接口落地", "把设计变成可运行的代码",
     ".claude/agents/代码工匠-红结.md", ".claude/commands/红结.md"],
    ["审计官", "旧影", "/旧影", "四级审计(目标对齐/红线执行/成本/技术健康)、问题追踪、修复路线", "告诉团队哪里出了问题，怎么修",
     ".claude/agents/审计官-旧影.md", ".claude/commands/旧影.md"],
]

for i, m in enumerate(members):
    r = row + 1 + i
    for c, v in enumerate(m, 1):
        style_data_cell(ws1, r, c, CENTER if c <= 3 else LEFT)
        ws1.cell(row=r, column=c, value=v)
    # Alternate row color
    if i % 2 == 0:
        for c in range(1, len(headers) + 1):
            ws1.cell(row=r, column=c).fill = LIGHT_BG

auto_width(ws1)

# ============ Sheet 2: 详细职能 ============
ws2 = wb.create_sheet("详细职能")

ws2.merge_cells("A1:C1")
ws2.cell(row=1, column=1, value="各角色详细职能").font = TITLE_FONT
ws2.cell(row=1, column=1).alignment = CENTER

detail_data = [
    ("阿黑 — 项目总监", "金融线委托腰子全权负责；工程线直接管理；跨线协调和项目全局运转；主动运营提醒"),
    ("腰子 — 金融业务负责人", "金融分析决策核心；善用团队：宏观→山猫、数据→玉夜、风险→流金、策略→青山；六维评分+综合判断；对金融判断质量负责"),
    ("山猫 — 宏观巡检官", "每日盘前宏观简报（政策+数据+事件+情绪）；经济日历前瞻；政策信号解读；为腰子提供宏观环境判断"),
    ("玉夜 — 数据监理", "三大API连通性巡检；缓存新鲜度审计、数据完整性验证；1+2架构合规检查；为腰子保障数据质量"),
    ("流金 — 风控官", "持仓集中度、相关性、流动性、事件风险；策略过拟合检测、回撤监控、交易纪律审查；为腰子设定风险边界"),
    ("青山 — 策略研究员", "因子IC/ICIR/分层收益/衰退趋势；策略优化提案；新因子探索、规律挖掘；为腰子提供策略优化提案"),
    ("情墨 — 系统架构师", "系统模块设计与边界划分、模块间接口契约定义；技术选型评估；数据流设计；重构决策；不写代码，出设计文档和架构图"),
    ("千光 — 构建工程师(CI)", "CI/CD流水线设计；自动化调度编排；定时任务管理；代码生成方案；不写代码，出流水线设计和调度方案"),
    ("红枫 — 部署工程师(环境)", "环境配置管理；部署策略设计；运行监控方案；备份与恢复；不写代码，出部署方案和环境诊断"),
    ("新安 — 质量工程师(测试)", "测试策略设计；回归测试用例；变更影响分析；代码规范审查；不写代码，出测试方案和审查报告"),
    ("红结 — 代码工匠", "PowerShell/Python编码实现；接口落地；代码重构执行；按情墨设计和千光流水线规范实现"),
    ("旧影 — 审计官", "独立审计，直接对阿黑负责；四级审计体系(目标对齐/红线执行/成本/技术健康)；发现问题、给出修复指令、追踪修复状态"),
]

h2 = ["角色", "详细职能", "硬边界"]
row2 = 3
for c, h in enumerate(h2, 1):
    ws2.cell(row=row2, column=c, value=h)
style_header_row(ws2, row2, len(h2))

boundaries = [
    "不自己分析，不越界替角色决策",
    "不写代码、不编造数据、不预测精确价位",
    "不写代码、不编造政策信息、不提供个股买卖建议",
    "不修代码，只给诊断报告和修复指令",
    "不写代码、不提供个股买卖建议",
    "不跑代码，设计方案给Claude执行回测",
    "不写代码、不执行部署或运维、不提供金融分析",
    "不写代码、不管理服务器环境、不做架构设计",
    "不写代码、不设计架构、不构建流水线",
    "不写代码、不修bug、不设计架构或流水线",
    "不设计架构、不出测试方案、只按规范写代码",
    "只审计不执行：发现问题、给出修复指令、追踪修复状态，不自己动手修代码",
]

for i, (name, detail) in enumerate(detail_data):
    r = row2 + 1 + i
    style_data_cell(ws2, r, 1, LEFT)
    style_data_cell(ws2, r, 2, LEFT_TOP)
    style_data_cell(ws2, r, 3, LEFT_TOP)
    ws2.cell(row=r, column=1, value=name)
    ws2.cell(row=r, column=2, value=detail)
    ws2.cell(row=r, column=3, value=boundaries[i])
    if i % 2 == 0:
        for c in range(1, 4):
            ws2.cell(row=r, column=c).fill = LIGHT_BG

# Column widths
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 65
ws2.column_dimensions['C'].width = 35

# ============ Sheet 3: 协作规则 ============
ws3 = wb.create_sheet("协作规则")

ws3.merge_cells("A1:B1")
ws3.cell(row=1, column=1, value="团队协作规则").font = TITLE_FONT
ws3.cell(row=1, column=1).alignment = CENTER

# Information flow chain
ws3.merge_cells("A3:B3")
ws3.cell(row=3, column=1, value="信息流").font = SUBTITLE_FONT
ws3.merge_cells("A4:B4")
ws3.cell(row=4, column=1, value="山猫+玉夜+流金+青山 → 腰子(综合判断) → 投资决策").font = Font(name="Microsoft YaHei", color="16213E", size=10, bold=True)
ws3.cell(row=4, column=1).alignment = CENTER
ws3.cell(row=4, column=1).fill = LIGHT_BG

# Coverage chain
ws3.merge_cells("A6:B6")
ws3.cell(row=6, column=1, value="全链路覆盖").font = SUBTITLE_FONT

chain_headers = ["环节", "角色 · 覆盖内容"]
r = 7
for c, h in enumerate(chain_headers, 1):
    ws3.cell(row=r, column=c, value=h)
style_header_row(ws3, r, 2)

chain_data = [
    ("宏观环境", "山猫 — 政策信号 / 经济日历 / 市场情绪 → 输出给腰子"),
    ("数据管线", "玉夜 — API巡检 / 缓存审计 / 完整性验证 → 输出给腰子"),
    ("分析决策(核心)", "腰子 — 六维评分 / 综合判断 / 最终决策 ← 接收四方支撑"),
    ("风险审计", "流金 — 集中度 / 过拟合 / 纪律审查 → 输出给腰子"),
    ("策略进化", "青山 — 因子IC / 参数优化 / 规律挖掘 → 输出给腰子"),
    ("系统设计", "情墨 — 模块划分 / 接口契约 / 技术选型"),
    ("构建流水线", "千光 — CI/CD / 自动化调度 / 定时任务"),
    ("部署环境", "红枫 — 环境配置 / 依赖管理 / 版本发布"),
    ("质量验证", "新安 — 测试策略 / 回归验证 / 变更分析"),
    ("代码实现", "红结 — PowerShell/Python编码、脚本优化"),
    ("独立审计", "旧影 — 四级审计体系 / 问题追踪 / 修复路线 → 直接对阿黑负责"),
]

for i, (label, detail) in enumerate(chain_data):
    r2 = r + 1 + i
    style_data_cell(ws3, r2, 1, CENTER)
    style_data_cell(ws3, r2, 2, LEFT)
    ws3.cell(row=r2, column=1, value=label)
    ws3.cell(row=r2, column=2, value=detail)
    if i % 2 == 0:
        for c in range(1, 3):
            ws3.cell(row=r2, column=c).fill = LIGHT_BG

# Rules
rules_start = r + 7
ws3.merge_cells(f"A{rules_start}:B{rules_start}")
ws3.cell(row=rules_start, column=1, value="协作纪律").font = SUBTITLE_FONT

rules = [
    ("腰子是金融线核心", "所有金融分析任务由腰子主导，协调四位专家，综合各方输出做最终投资判断"),
    ("支撑角色向腰子汇报", "山猫/玉夜/流金/青山的输出统一给腰子；腰子不在时，阿黑代为接收"),
    ("分工明确，互不越界", "腰子不做风控细节，流金不挖因子，青山不看单票，山猫不做个股；红结只写代码"),
    ("金融信息流", "山猫+玉夜+流金+青山 → 腰子 → 投资决策"),
    ("工程信息流", "情墨设计 → 千光构建 → 红枫部署 → 新安验证 → 红结实现"),
    ("跨团队协作", "腰子提金融需求 → 阿黑分派工程团队 → 工程团队实施 → 腰子验证金融结果"),
    ("争议解决", "金融团队内→腰子判断；工程团队内→阿黑判断；跨线→阿黑协调，用户最终决策"),
    ("红线共守", "所有角色共同遵守规则红线最新版本"),
    ("审计独立", "旧影独立于金融线和工程线，直接对阿黑负责；审计结论不受被审计方影响；任何人不得阻止或修改审计发现"),
]

for i, (rule, desc) in enumerate(rules):
    r3 = rules_start + 1 + i
    style_data_cell(ws3, r3, 1, LEFT)
    style_data_cell(ws3, r3, 2, LEFT)
    ws3.cell(row=r3, column=1, value=rule)
    ws3.cell(row=r3, column=2, value=desc)
    ws3.cell(row=r3, column=1).font = Font(name="Microsoft YaHei", bold=True, color="1A1A2E", size=10)
    if i % 2 == 0:
        for c in range(1, 3):
            ws3.cell(row=r3, column=c).fill = LIGHT_BG

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 65

# ============ Sheet 4: 文件索引 ============
ws4 = wb.create_sheet("文件索引")

ws4.merge_cells("A1:C1")
ws4.cell(row=1, column=1, value="相关文件索引").font = TITLE_FONT
ws4.cell(row=1, column=1).alignment = CENTER

idx_headers = ["类型", "路径", "说明"]
r = 3
for c, h in enumerate(idx_headers, 1):
    ws4.cell(row=r, column=c, value=h)
style_header_row(ws4, r, len(idx_headers))

files = [
    ("本名册(v1.8 .md)", "项目成员/团队名册_v1.8.md", "Markdown 源文件（当前版本）"),
    ("本名册(v1.8 .xlsx)", "项目成员/团队名册_v1.8.xlsx", "Excel 版本（当前版本）"),
    ("本名册(v1.5 归档)", "项目成员/团队名册_v1.5.md/.xlsx", "历史版本归档"),
    ("", "", ""),
    ("金融团队角色定义", "", ""),
    ("阿黑角色定义", ".claude/agents/项目总监-阿黑.md", "项目总监完整人设 (v1.1)"),
    ("腰子角色定义", ".claude/agents/金融专家-腰子.md", "金融业务负责人完整人设 (v2.1)"),
    ("流金角色定义", ".claude/agents/风控官-流金.md", "风控官完整人设 (v2.1)"),
    ("玉夜角色定义", ".claude/agents/数据监理-玉夜.md", "数据监理完整人设 (v2.1)"),
    ("青山角色定义", ".claude/agents/策略研究员-青山.md", "策略研究员完整人设 (v2.1)"),
    ("山猫角色定义", ".claude/agents/宏观巡检-山猫.md", "宏观巡检完整人设 (v2.1)"),
    ("", "", ""),
    ("工程团队角色定义", "", ""),
    ("情墨角色定义", ".claude/agents/系统架构师-情墨.md", "系统架构师完整人设 (v2.0)"),
    ("千光角色定义", ".claude/agents/构建工程师-千光.md", "构建工程师(CI)完整人设 (v2.0)"),
    ("红枫角色定义", ".claude/agents/部署工程师-红枫.md", "部署工程师(环境)完整人设 (v2.0)"),
    ("新安角色定义", ".claude/agents/质量工程师-新安.md", "质量工程师(测试)完整人设 (v2.0)"),
    ("红结角色定义", ".claude/agents/代码工匠-红结.md", "代码工匠完整人设 (v1.0)"),
    ("", "", ""),
    ("独立审计角色定义", "", ""),
    ("旧影角色定义", ".claude/agents/审计官-旧影.md", "审计官完整人设 (v1.0)"),
    ("", "", ""),
    ("金融团队协作协议", ".claude/agents/金融团队-协作协议.md", "金融团队内部协作规范 (v1.0)"),
    ("金融团队学习计划", ".claude/agents/金融团队-学习计划.md", "支撑角色定向培训 (v1.0)"),
    ("", "", ""),
    ("知识库目录", "", ""),
    ("腰子知识库 (15+1文件)", ".claude/agents/腰子-知识库/", "金融专家知识体系"),
    ("流金知识库 (6+1文件)", ".claude/agents/流金-知识库/", "风控知识体系"),
    ("青山知识库 (5+1文件)", ".claude/agents/青山-知识库/", "策略研究知识体系"),
    ("山猫知识库 (6+1文件)", ".claude/agents/山猫-知识库/", "宏观巡检知识体系"),
    ("玉夜知识库 (5+1文件)", ".claude/agents/玉夜-知识库/", "数据监理知识体系"),
    ("情墨知识库 (5+1文件)", ".claude/agents/情墨-知识库/", "系统架构知识体系"),
    ("千光知识库 (4+1文件)", ".claude/agents/千光-知识库/", "构建工程知识体系"),
    ("红枫知识库 (4+1文件)", ".claude/agents/红枫-知识库/", "部署工程知识体系"),
    ("新安知识库 (5+1文件)", ".claude/agents/新安-知识库/", "质量工程知识体系"),
    ("红结知识库 (5+1文件)", ".claude/agents/红结-知识库/", "代码实现知识体系"),
    ("旧影知识库 (3文件)", ".claude/agents/旧影-知识库/", "审计知识体系"),
    ("", "", ""),
    ("召唤命令 ×11", ".claude/commands/腰子.md 等", "角色召唤触发器"),
    ("名册生成工具", "代码文件/tools/generate_roster_xlsx.py", "自动生成团队名册.xlsx"),
    ("规则红线", "规则红线/分析的规则红线--Claude_v1.9.md", "项目最高优先级规则"),
]

for i, (ftype, fpath, fdesc) in enumerate(files):
    r2 = r + 1 + i
    style_data_cell(ws4, r2, 1, CENTER)
    style_data_cell(ws4, r2, 2, LEFT)
    style_data_cell(ws4, r2, 3, LEFT)
    ws4.cell(row=r2, column=1, value=ftype)
    ws4.cell(row=r2, column=2, value=fpath)
    ws4.cell(row=r2, column=3, value=fdesc)
    if i % 2 == 0:
        for c in range(1, 4):
            ws4.cell(row=r2, column=c).fill = LIGHT_BG

auto_width(ws4)

# Freeze panes for all sheets
ws1.freeze_panes = "A10"
ws2.freeze_panes = "A4"
ws3.freeze_panes = "A3"
ws4.freeze_panes = "A4"

output_path = r"c:\Users\34269\Documents\Claude\股票分析\项目成员\团队名册_v1.8.xlsx"
wb.save(output_path)
print(f"[OK] 团队名册.xlsx generated at {output_path}")
